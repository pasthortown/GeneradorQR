import qrcode
import openpyxl
import os

archivo_excel = 'codigos.xlsx'
wb = openpyxl.load_workbook(archivo_excel)
hoja = wb['Hoja1']
codes = [hoja.cell(row=i, column=1).value for i in range(1, hoja.max_row + 1)]

def generar_qr(valor, nombre_archivo="codigo_qr.png"):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(valor)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(nombre_archivo)

carpeta_codigos = 'codigos'
if not os.path.exists(carpeta_codigos):
    os.makedirs(carpeta_codigos)

for codigo in codes:
    valor_a_codificar = codigo
    nombre_del_archivo = os.path.join(carpeta_codigos, f"{codigo}.png")
    generar_qr(valor_a_codificar, nombre_del_archivo)

print("Se han generado los códigos QR respectivos.")
